import sys
if sys.platform == 'win32':
    import asyncio
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

import streamlit as st
from playwright.async_api import async_playwright, Page, TimeoutError
from playwright.sync_api import sync_playwright
from bs4 import BeautifulSoup
import json
import io
import logging
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import re
from collections import defaultdict
import asyncio
from datetime import datetime

# ------------------------- LLM AGENT SETUP (Updated for LangChain 2024) -------------------------
from langchain_groq import ChatGroq
from langchain_core.prompts import ChatPromptTemplate
from langchain.chains import LLMChain

GROQ_API_KEY = "your_groq_api_key_here"
MODEL_NAME = "meta-llama/llama-4-scout-17b-16e-instruct"

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def get_chat_llm():
    return ChatGroq(
        temperature=0.08,
        groq_api_key=GROQ_API_KEY,
        model=MODEL_NAME,
        max_tokens=2048,
    )

async def setup_browser() -> tuple[Page, any, any]:
    """Set up a Playwright browser instance with custom configurations."""
    try:
        playwright = await async_playwright().start()
        browser = await playwright.chromium.launch(
            headless=False,
            args=["--no-sandbox", "--disable-gpu", "--window-size=1920,1080"],
        )
        context = await browser.new_context()
        page = await context.new_page()
        logger.info("Browser initialized successfully")
        return page, browser, playwright
    except Exception as e:
        logger.error(f"Browser setup failed: {str(e)}")
        return None, None, None

async def execute_action(page, action_data):
    """Execute a browser action based on the action type and parameters."""
    if not page:
        logger.error("No browser page available")
        return None
    action_type = action_data.get("action")
    params = {k: v for k, v in action_data.items() if k not in ["action", "task_id", "description"]}
    
    try:
        if action_type == "open_website":
            await page.goto(params.get("url", ""), timeout=15000)
            await page.wait_for_load_state("networkidle", timeout=15000)
            logger.info(f"Navigated to {params.get('url', 'N/A')}")
            return True
        elif action_type == "click_element":
            selector = params.get("selector", "")
            await page.wait_for_selector(selector, timeout=10000)
            element = await page.query_selector(selector)
            if element and await element.is_visible():
                await element.click()
                await asyncio.sleep(2)  # Wait for page stability
                logger.info(f"Clicked on selector {selector}")
                return True
            else:
                logger.error(f"Element not found or not visible: {selector}")
                return False
        elif action_type == "fill_input":
            selector = params.get("selector", "")
            value = params.get("value", "")
            await page.wait_for_selector(selector, timeout=5000)
            await page.fill(selector, value)
            logger.info(f"Filled input {selector} with {value}")
            return True
        elif action_type == "verify_title":
            content = await page.title()
            expected = params.get("value", "")
            result = expected.lower() in content.lower()
            logger.info(f"Verified title: {content}, Expected: {expected}, Result: {result}")
            return result
        elif action_type == "verify_description":
            try:
                content = await page.evaluate('document.querySelector("meta[name=\'description\']")?.content || ""')
                expected = params.get("value", "")
                result = expected.lower() in content.lower()
                logger.info(f"Verified description: {content}, Expected: {expected}, Result: {result}")
                return result
            except:
                logger.warning("Could not find meta description")
                return False
        elif action_type == "verify_url":
            current_url = page.url
            expected = params.get("value", "")
            result = expected.lower() in current_url.lower()
            logger.info(f"Verified URL: {current_url}, Expected: {expected}, Result: {result}")
            return result
        elif action_type == "verify_url_contains":
            current_url = page.url
            expected = params.get("value", "")
            result = expected.lower() in current_url.lower()
            logger.info(f"Verified URL contains: {current_url}, Expected: {expected}, Result: {result}")
            return result
        elif action_type == "verify_text":
            selector = params.get("selector", "body")
            expected = params.get("value", "")
            try:
                await page.wait_for_selector(selector, timeout=5000)
                content = await page.text_content(selector)
                result = expected.lower() in (content.lower() if content else "")
                logger.info(f"Verified text: Found '{expected}' in content: {result}")
                return result
            except:
                logger.error(f"Could not verify text in selector: {selector}")
                return False
        elif action_type == "verify_element_exists":
            selector = params.get("selector", "")
            try:
                await page.wait_for_selector(selector, timeout=5000)
                element = await page.query_selector(selector)
                result = element is not None and await element.is_visible()
                logger.info(f"Verified element exists: {selector}, Result: {result}")
                return result
            except:
                logger.error(f"Element does not exist: {selector}")
                return False
        elif action_type == "verify_error_or_redirect":
            content = await page.content()
            result = "error" in content.lower() or "404" in content.lower() or "not found" in content.lower()
            logger.info(f"Verified error page or redirect: Result: {result}")
            return result
        elif action_type == "verify_form_submission":
            # Check for success message or redirect after form submission
            try:
                await page.wait_for_load_state("networkidle", timeout=10000)
                content = await page.text_content("body")
                success_indicators = ["success", "thank you", "submitted", "sent", "received"]
                result = any(indicator in content.lower() for indicator in success_indicators)
                logger.info(f"Verified form submission: Result: {result}")
                return result
            except:
                return False
    except Exception as e:
        logger.error(f"Action failed: {action_type}, Error: {str(e)}")
        return False
    return True

def convert_test_case_to_browser_actions(test_case, base_url, scraped_data):
    """Convert a structured test case to browser actions."""
    actions = []
    test_steps = test_case.get('Test Steps', '')
    test_scenario = test_case.get('Test Scenario', '')
    
    # Always start by opening the website
    actions.append({
        "action": "open_website",
        "url": base_url,
        "description": "Navigate to the website"
    })
    
    # Parse test steps and convert to browser actions
    steps = test_steps.split('\n')
    
    for step in steps:
        step = step.strip()
        if not step:
            continue
            
        step_lower = step.lower()
        
        # Navigation actions
        if any(nav_word in step_lower for nav_word in ['navigate to', 'go to', 'visit', 'open']):
            if 'home' in step_lower:
                actions.append({
                    "action": "click_element",
                    "selector": "a:has-text('Home'), a[href='/'], a[href='#home']",
                    "description": f"Navigate to home page: {step}"
                })
            elif 'about' in step_lower:
                actions.append({
                    "action": "click_element",
                    "selector": "a:has-text('About'), a[href*='about']",
                    "description": f"Navigate to about page: {step}"
                })
            elif 'contact' in step_lower:
                actions.append({
                    "action": "click_element",
                    "selector": "a:has-text('Contact'), a[href*='contact']",
                    "description": f"Navigate to contact page: {step}"
                })
            elif 'features' in step_lower:
                actions.append({
                    "action": "click_element",
                    "selector": "a:has-text('Features'), a[href*='feature']",
                    "description": f"Navigate to features page: {step}"
                })
            elif 'products' in step_lower:
                actions.append({
                    "action": "click_element",
                    "selector": "a:has-text('Products'), a[href*='product']",
                    "description": f"Navigate to products page: {step}"
                })
                
        # Form filling actions
        elif any(form_word in step_lower for form_word in ['enter', 'fill', 'input', 'type']):
            if 'email' in step_lower:
                actions.append({
                    "action": "fill_input",
                    "selector": "input[type='email'], input[name*='email'], input[id*='email']",
                    "value": "test@example.com",
                    "description": f"Fill email field: {step}"
                })
            elif 'name' in step_lower:
                actions.append({
                    "action": "fill_input",
                    "selector": "input[name*='name'], input[id*='name']",
                    "value": "Test User",
                    "description": f"Fill name field: {step}"
                })
            elif 'message' in step_lower:
                actions.append({
                    "action": "fill_input",
                    "selector": "textarea, input[name*='message'], input[id*='message']",
                    "value": "This is a test message.",
                    "description": f"Fill message field: {step}"
                })
            elif 'password' in step_lower:
                actions.append({
                    "action": "fill_input",
                    "selector": "input[type='password'], input[name*='password']",
                    "value": "TestPassword123",
                    "description": f"Fill password field: {step}"
                })
                
        # Click actions
        elif any(click_word in step_lower for click_word in ['click', 'press', 'select']):
            if 'submit' in step_lower or 'send' in step_lower:
                actions.append({
                    "action": "click_element",
                    "selector": "button[type='submit'], input[type='submit'], button:has-text('Submit'), button:has-text('Send')",
                    "description": f"Click submit button: {step}"
                })
            elif 'login' in step_lower:
                actions.append({
                    "action": "click_element",
                    "selector": "button:has-text('Login'), input[value*='Login'], a:has-text('Login')",
                    "description": f"Click login button: {step}"
                })
            elif 'register' in step_lower or 'sign up' in step_lower:
                actions.append({
                    "action": "click_element",
                    "selector": "button:has-text('Register'), button:has-text('Sign Up'), a:has-text('Register')",
                    "description": f"Click register button: {step}"
                })
                
        # Verification actions
        elif any(verify_word in step_lower for verify_word in ['verify', 'check', 'confirm', 'validate']):
            if 'title' in step_lower:
                # Extract expected text from step
                expected_text = extract_expected_text(step)
                actions.append({
                    "action": "verify_title",
                    "value": expected_text or test_scenario,
                    "description": f"Verify page title: {step}"
                })
            elif 'url' in step_lower:
                expected_url = extract_expected_text(step)
                actions.append({
                    "action": "verify_url_contains",
                    "value": expected_url or test_scenario.lower().replace(' ', '-'),
                    "description": f"Verify URL: {step}"
                })
            elif 'text' in step_lower or 'content' in step_lower:
                expected_text = extract_expected_text(step)
                actions.append({
                    "action": "verify_text",
                    "selector": "body",
                    "value": expected_text or test_scenario,
                    "description": f"Verify page content: {step}"
                })
            elif 'error' in step_lower:
                actions.append({
                    "action": "verify_error_or_redirect",
                    "description": f"Verify error handling: {step}"
                })
    
    # Add verification based on expected result
    expected_result = test_case.get('Expected Result', '')
    if expected_result:
        if 'success' in expected_result.lower():
            actions.append({
                "action": "verify_form_submission",
                "description": f"Verify expected result: {expected_result}"
            })
        else:
            actions.append({
                "action": "verify_text",
                "selector": "body",
                "value": expected_result.split()[0] if expected_result else test_scenario,
                "description": f"Verify expected result: {expected_result}"
            })
    
    return actions

def extract_expected_text(step_text):
    """Extract expected text from a step description."""
    # Look for quoted text
    quotes = re.findall(r'"([^"]*)"', step_text)
    if quotes:
        return quotes[0]
    
    # Look for text after 'contains', 'shows', 'displays'
    patterns = [
        r'contains\s+["\']?([^"\']+)["\']?',
        r'shows\s+["\']?([^"\']+)["\']?',
        r'displays\s+["\']?([^"\']+)["\']?'
    ]
    
    for pattern in patterns:
        match = re.search(pattern, step_text, re.I)
        if match:
            return match.group(1).strip()
    
    return None

async def test_all_buttons(page: Page, main_url: str, scraped_data: dict = None) -> list[dict]:
    """Test all navigation menu items and verify page loads."""
    report = []
    
    # Get navigation items from scraped data if available
    nav_items = []
    if scraped_data and 'navigation' in scraped_data:
        nav_items = [nav['text'] for nav in scraped_data['navigation'] if nav['text']]
    
    # Fallback to common navigation items
    if not nav_items:
        nav_items = ["Naipunya AI Labs", "Home", "Features", "Capabilities", "Products", "About Us", "Contact"]
    
    try:
        logger.info(f"Navigating to {main_url}")
        await page.goto(main_url, wait_until="networkidle", timeout=15000)
        await page.wait_for_load_state("networkidle", timeout=15000)
        
        # Wait for navigation menu to load
        await page.wait_for_selector("nav, header, .navbar, .menu", timeout=10000)
        
        for item in nav_items:
            try:
                # Multiple selector strategies
                selectors = [
                    f"a:has-text('{item}')",
                    f"a[href*='{item.lower().replace(' ', '-')}']",
                    f"a[href*='{item.lower().replace(' ', '')}']",
                    f"*[role='menuitem']:has-text('{item}')",
                    f"li:has-text('{item}') a",
                    f"nav a:has-text('{item}')"
                ]
                
                element = None
                used_selector = None
                
                for selector in selectors:
                    try:
                        element = await page.wait_for_selector(selector, timeout=2000)
                        if element and await element.is_visible():
                            used_selector = selector
                            break
                    except:
                        continue
                
                if not element:
                    logger.error(f"Navigation item {item} not found with any selector")
                    report.append({
                        "button_label": item,
                        "clickable": False,
                        "loaded_successfully": False,
                        "errors": "Element not found"
                    })
                    continue
                
                clickable = await element.is_enabled() and await element.is_visible()
                if not clickable:
                    logger.warning(f"Navigation item {item} not clickable or visible")
                    report.append({
                        "button_label": item,
                        "clickable": False,
                        "loaded_successfully": False,
                        "errors": "Element is disabled or not visible"
                    })
                    continue
                
                before_url = page.url
                logger.info(f"Clicking {item} with selector: {used_selector}")
                await element.click()
                await asyncio.sleep(3)  # Wait for page stability
                await page.wait_for_load_state("networkidle", timeout=15000)
                
                after_url = page.url
                error = None
                loaded = True
                
                # Verify page content
                page_content = (await page.inner_text("body")).lower()
                page_title = (await page.title()).lower()
                
                # Specific verification for each page
                expected_content = item.lower().replace(' ', '')
                if expected_content not in page_content and expected_content not in page_title:
                    # More flexible content verification
                    item_words = item.lower().split()
                    found_words = sum(1 for word in item_words if word in page_content or word in page_title)
                    if found_words == 0:
                        loaded = False
                        error = f"Expected content related to '{item}' not found in page"
                
                # Navigate back to main page if URL changed
                if before_url != after_url:
                    logger.info(f"Navigating back to {main_url}")
                    await page.goto(main_url, wait_until="networkidle", timeout=15000)
                    await page.wait_for_load_state("networkidle", timeout=15000)
                
                report.append({
                    "button_label": item,
                    "clickable": True,
                    "loaded_successfully": loaded,
                    "errors": error,
                    "selector_used": used_selector
                })
                logger.info(f"Tested {item}: Loaded={loaded}, Error={error}")
                
            except TimeoutError as e:
                logger.error(f"Timeout while testing {item}: {str(e)}")
                report.append({
                    "button_label": item,
                    "clickable": False,
                    "loaded_successfully": False,
                    "errors": f"Timeout: {str(e)}"
                })
            except Exception as e:
                logger.error(f"Failed to test {item}: {str(e)}")
                report.append({
                    "button_label": item,
                    "clickable": False,
                    "loaded_successfully": False,
                    "errors": str(e)
                })
                
    except Exception as e:
        logger.error(f"Button testing failed: {str(e)}")
        report.append({
            "button_label": "N/A",
            "clickable": False,
            "loaded_successfully": False,
            "errors": f"Failed to process buttons: {str(e)}"
        })
    
    return report

async def run_browser_test_agent(test_cases_data, base_url, scraped_data):
    """Run the browser test agent with test cases from structured data."""
    page, browser, playwright = await setup_browser()
    if not page:
        logger.error("Aborting due to browser initialization failure")
        return None, None, True

    test_results = []
    button_report = None
    error_occurred = False

    for test_case in test_cases_data:
        try:
            test_id = test_case.get('Test Case ID', 'UNKNOWN')
            logger.info(f"Executing test case: {test_id}")
            
            # Convert test case to browser actions
            actions = convert_test_case_to_browser_actions(test_case, base_url, scraped_data)
            
            step_results = []
            for action in actions:
                try:
                    result = await execute_action(page, action)
                    step_results.append({
                        "action": action,
                        "result": result,
                        "description": action.get("description", "")
                    })
                    logger.info(f"Action result: {action['action']} = {result}")
                except Exception as e:
                    logger.error(f"Action execution failed: {str(e)}")
                    step_results.append({
                        "action": action,
                        "result": False,
                        "error": str(e),
                        "description": action.get("description", "")
                    })
                    error_occurred = True
            
            test_results.append({
                "test_id": test_id,
                "test_scenario": test_case.get('Test Scenario', ''),
                "steps": step_results,
                "overall_status": "PASS" if all(step.get("result", False) for step in step_results) else "FAIL"
            })
            
        except Exception as e:
            logger.error(f"Test case {test_case.get('Test Case ID', 'UNKNOWN')} failed: {str(e)}")
            error_occurred = True
            test_results.append({
                "test_id": test_case.get('Test Case ID', 'UNKNOWN'),
                "error": str(e),
                "overall_status": "ERROR"
            })

    # Test all buttons on the website
    logger.info("Starting button testing...")
    button_report = await test_all_buttons(page, base_url, scraped_data)

    # Cleanup
    try:
        await page.close()
        await browser.close()
        await playwright.stop()
    except Exception as e:
        logger.error(f"Cleanup failed: {str(e)}")
    logger.info("Browser and context closed successfully")

    return test_results, button_report, error_occurred

def scrape_website_sync(url):
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(url, wait_until="networkidle")
        html = page.content()
        browser.close()
    soup = BeautifulSoup(html, "lxml")

    navigation = []
    for nav in soup.find_all(['nav', 'ul', 'menu']):
        for link in nav.find_all('a', href=True):
            text = link.get_text(strip=True)
            href = link.get("href")
            if text:
                navigation.append({'text': text, 'href': href})
    seen = set()
    navigation = [d for d in navigation if not (d['href'], d['text']) in seen and not seen.add((d['href'], d['text']))]

    page_info = {}
    title_tag = soup.find('title')
    if title_tag: page_info["title"] = title_tag.get_text(strip=True)
    meta_desc = soup.find('meta', attrs={'name': 'description'})
    if meta_desc: page_info['description'] = meta_desc.get('content','')
    page_info["headings"] = [h.get_text(strip=True) for h in soup.find_all(['h1','h2','h3'])]

    forms = []
    for f in soup.find_all('form'):
        form_info = {"action": f.get('action'), "method": f.get('method','get'), "fields": [], "buttons": []}
        for e in f.find_all(['input','textarea','select']):
            input_type = e.get('type', e.name)
            name = e.get('name')
            label = None
            if e.has_attr('id'):
                label_tag = soup.find('label', attrs={'for': e['id']})
                if label_tag: label = label_tag.get_text(strip=True)
            if not label:
                parent_label = e.find_parent('label')
                if parent_label: label = parent_label.get_text(strip=True)
            form_info['fields'].append({'name': name, 'type': input_type, 'label': label})
        for b in f.find_all(['button','input']):
            if b.name == 'button' or (b.name == 'input' and b.get('type') in ('submit','button')):
                btext = b.get_text(strip=True) if b.name == 'button' else b.get('value', '')
                form_info['buttons'].append({'text': btext, 'type': b.get('type', 'submit')})
        forms.append(form_info)

    extra_buttons = []
    for b in soup.find_all(['button','input']):
        if not b.find_parent('form'):
            if b.name == 'button' or (b.name == 'input' and b.get('type') in ('submit', 'button')):
                btext = b.get_text(strip=True) if b.name == 'button' else b.get('value','')
                extra_buttons.append({'text': btext, 'type': b.get('type','button')})

    interact_links = []
    for a in soup.find_all('a', href=True):
        if a.get_text(strip=True) and not a.find_parent(['nav','menu','form','ul']):
            interact_links.append({'text': a.get_text(strip=True), 'href': a['href']})

    alerts = []
    for el in soup.find_all(attrs={"role": "alert"}):
        alerts.append(el.get_text(strip=True))
    for el in soup.find_all(class_=lambda x: x and 'error' in x):
        alerts.append(el.get_text(strip=True))
    for el in soup.find_all(attrs={"aria-live": True}):
        alerts.append(el.get_text(strip=True))
    alerts = list(set([a for a in alerts if a]))

    output = {
        'navigation': navigation,
        'page_info': page_info,
        'forms': forms,
        'extra_buttons': extra_buttons,
        'interactive_links': interact_links,
        'alert_messages': alerts,
        'url': url
    }
    return output

# ------------------------- PROMPT GENERATORS -------------------------
def get_user_story_prompt():
    return """
You are a professional Agile Product Owner.
Given a structured extraction from a website (as JSON) showing its navigation, headings, forms, buttons, links, and error messages, write a precise, numbered list of USER STORIES for a "Guest User" (not logged in).
Guidelines:
- Use the format: "As a Guest, I want to [do something] so that [reason/benefit/goal]."
- Cover each unique feature, action, or content area observed (navigation, forms, about/contact, etc.).
- For login/registration forms, describe the main intent.
- For About or information sections, frame stories about discovering company/site info.
- For buttons/links that perform actions, describe the desired action/result.
- For error or alert messages, express needs for clear/understandable feedback.
- Do NOT speculate features that are not evident.
- Write for clarity and completeness, in professional Agile style.
Here's the extracted info:
{structured_info}
"""

def get_test_plan_prompt():
    return """
You are a senior QA Test Lead.
Given the structured extraction of a website (JSON) with navigation, headings, forms, buttons, links, and error/alert messages, create a concise, professional "Test Plan" section for a test document for Guest User.
Guidelines:
- Summarize main objectives for guest user testing.
- List key features/areas under test.
- Outline test scope (what is covered, what isn't, e.g. only guest user, not admin).
- List limitations/assumptions if any.
- Write in formal language in proper test plan section format (max 200 words).
Here's the extracted info:
{structured_info}
"""

def get_test_case_prompt():
    return """
You are a senior QA engineer.
Given the structured extraction of a website (JSON) with navigation, headings, forms, buttons, links, and error/alert messages, generate a structured list of end-to-end TEST CASES a "Guest User" can perform, mapped to the apparent features and functionalities (navigation, forms, about, etc.)
Guidelines:
- For each unique feature (e.g. login form, contact us, about, etc.), enumerate the "happy path" and at least one "negative path" (if errors/alerts found).
- Use clear step-by-step actions and expected results.
- Use this format for each test:
Test Case #: [ID]
Title: [short descriptive name]
Steps:
1. [action step]
...
Expected Result: [result]
- Focus ONLY on features actually present in the extracted info. Do not invent details.
Here is the extracted info:
{structured_info}
"""

def get_test_data_prompt():
    return """
You are a QA Test Data Generator. Given this test case, generate realistic dummy test data for each input, in the format: 
"Field: Value"
for each input field mentioned in the steps (e.g. name, email, etc.).
If test case does not need input, return "N/A"
Here is the test case:
{test_case}
"""

# ----------------------- LLM Agent Wrappers --------------------------
def call_llm_agent(prompt_template, input_key="structured_info", input_value=None):
    prompt = ChatPromptTemplate.from_template(prompt_template)
    llm = get_chat_llm()
    chain = LLMChain(llm=llm, prompt=prompt)
    resp = chain.invoke({input_key: input_value})
    return resp['text']

def generate_user_stories(site_data):
    json_data = json.dumps(site_data, indent=2)
    output = call_llm_agent(get_user_story_prompt(), "structured_info", json_data)
    return output

def generate_test_plan(site_data):
    json_data = json.dumps(site_data, indent=2)
    output = call_llm_agent(get_test_plan_prompt(), "structured_info", json_data)
    return output

def generate_test_cases(site_data):
    json_data = json.dumps(site_data, indent=2)
    output = call_llm_agent(get_test_case_prompt(), "structured_info", json_data)
    return output

def generate_test_data_for_case(testcase_string):
    out = call_llm_agent(get_test_data_prompt(), "test_case", testcase_string)
    return out.strip()

# ----------------------- Excel Formatting Utilities -----------------
def style_user_stories_excel(parsed_stories):
    wb = Workbook()
    ws = wb.active
    ws.title = "User Stories"
    hdr_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    headers = ["ID", "User Story"]
    ws.append(headers)
    for col, h in enumerate(headers,1):
        cell = ws.cell(1,col)
        cell.fill = hdr_fill
        cell.font = bold
        cell.alignment = wrap
        ws.column_dimensions[get_column_letter(col)].width = 40 if col==2 else 12
    for row in parsed_stories:
        ws.append([row["ID"], row["User Story"]])
    for row in ws.iter_rows(min_row=2,max_col=2):
        for cell in row:
            cell.font = Font(size=11)
            cell.alignment = wrap
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def style_test_case_excel(parsed_cases):
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    hdr_fill = PatternFill(start_color="FFE5B4", end_color="FFE5B4", fill_type="solid")
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    headers = ["Test Case ID", "User Story", "Test Scenario", "Preconditions", "Test Steps", "Test Data", "Expected Result", "Actual Result", "Status", "Priority"]
    ws.append(headers)
    for col, h in enumerate(headers,1):
        cell = ws.cell(1,col)
        cell.fill = hdr_fill
        cell.font = bold
        cell.alignment = wrap
        ws.column_dimensions[get_column_letter(col)].width = 36 if col!=1 else 18
    for tc in parsed_cases:
        ws.append([tc.get(col,"") for col in headers])
    for row in ws.iter_rows(min_row=2,max_col=len(headers)):
        for cell in row:
            cell.font = Font(size=11)
            cell.alignment = wrap
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def style_button_report_excel(button_report):
    wb = Workbook()
    ws = wb.active
    ws.title = "Button Test Report"
    hdr_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    headers = ["Button Label", "Clickable", "Loaded Successfully", "Errors", "Selector Used"]
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.fill = hdr_fill
        cell.font = bold
        cell.alignment = wrap
        ws.column_dimensions[get_column_letter(col)].width = 40 if col == 1 else 20
    for report in button_report:
        ws.append([
            report.get("button_label", ""),
            str(report.get("clickable", False)),
            str(report.get("loaded_successfully", False)),
            report.get("errors", ""),
            report.get("selector_used", "")
        ])
    for row in ws.iter_rows(min_row=2, max_col=len(headers)):
        for cell in row:
            cell.font = Font(size=11)
            cell.alignment = wrap
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def style_test_execution_report_excel(report_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Execution Results"
    headers = ["Test Case ID", "Test Scenario", "Test Steps", "Expected Result", "Actual Result", "Status", "Execution Timestamp"]
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.fill = PatternFill(start_color="FFE5B4", end_color="FFE5B4", fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions[get_column_letter(col)].width = 30
    for row_data in report_data:
        ws.append([row_data.get(h, "") for h in headers])
    for row in ws.iter_rows(min_row=2, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ---------------------- Markdown Parsing Utilities --------------------
def parse_markdown_user_stories(output_str):
    lines = output_str.splitlines()
    stories = []
    idx = 1
    for l in lines:
        l = l.strip()
        l = re.sub(r'^[\*\-\d\. ]+','',l)
        l = re.sub(r'^[*_]{1,3}|[_*]{1,3}$','',l)
        if l.lower().startswith("as a guest") and len(l) > 20:
            uid = f"US-{idx:03d}"
            stories.append({'ID': uid, 'User Story': l.strip()})
            idx += 1
    return stories

def clean_markdown(text):
    return re.sub(r'[*_]{1,2}', '', text).strip()

def parse_markdown_test_cases(markdown):
    test_cases = []
    blocks = re.split(r"Test Case #:", markdown)
    for block in blocks:
        block = block.strip()
        if not block or not re.search(r"Title:", block):
            continue
        title_match = re.search(r"Title:\s*(.+)", block)
        title = title_match.group(1).strip() if title_match else ""
        steps = []
        steps_found = False
        for line in block.splitlines():
            l = line.strip()
            if re.match(r'^\d+\.\s', l) and not l.lower().startswith('expected result'):
                steps_found = True
                steps.append(clean_markdown(l))
            elif steps_found and ('expected result:' in l.lower()):
                break
            elif steps_found and l and not l.lower().startswith('expected result:'):
                steps.append(clean_markdown(l))
        ex_match = re.search(r"Expected Result\s*:*(.+)", block, re.I)
        expect = clean_markdown(ex_match.group(1).strip()) if ex_match else ""
        if not expect:
            for line in reversed(block.splitlines()):
                if "expected" in line.lower():
                    expect = clean_markdown(line.split(":", 1)[-1])
                    break
        test_cases.append({
            'Test Case ID': '',
            'User Story': '',
            'Test Scenario': clean_markdown(title),
            'Preconditions': "",
            'Test Steps': "\n".join(steps),
            'Test Data': "",
            'Expected Result': expect,
            'Actual Result': "",
            'Status': "",
            'Priority': "Medium"
        })
    return test_cases

def assign_test_case_ids(test_cases):
    story_groups = defaultdict(list)
    for tc in test_cases:
        story = tc.get('User Story') or "US-000"
        story_groups[story].append(tc)
    for story, group in story_groups.items():
        us_num = re.sub(r'\D','', story) or '000'
        for i, tc in enumerate(group, 1):
            tc['Test Case ID'] = f"TC-{us_num}-{i:02d}"
    return test_cases

# ---------------------- Streamlit UI with Tabs/Session State ----------------------
st.set_page_config(page_title="🤖 Agile User Story & Test Case Agent", layout="wide")
st.title("🧩 Website User Stories & Test Cases Generator (Guest User, AI Agent)")

TABS = [
    "Website Scraping",
    "🧩 User Story Creation Agent",
    "🧪 Test Plan Creation Agent", 
    "🧪 Test Case / Test Data Agent",
    "🤖 Browser Test Agent"
]
tab = st.tabs(TABS)

# ------------- TAB 1: SCRAPING -------------
with tab[0]:
    url = st.text_input("Enter website URL (must include https://):", "https://naipunya.ai/")
    if st.button("Scrape and Analyze"):
        with st.spinner("Scraping website and analyzing structure..."):
            try:
                scraped_data = scrape_website_sync(url)
                st.session_state['scraped_data'] = scraped_data
                st.success("Scraping complete! Use the expanders below to inspect.")
            except Exception as e:
                st.error(f"Scraping error: {str(e)}")
                st.stop()
    if 'scraped_data' in st.session_state:
        data = st.session_state['scraped_data']
        st.write(f"**URL:** {data['url']}")
        with st.expander("Navigation Menu"):
            st.json(data['navigation'])
        with st.expander("Page Info (headings/meta/title)"):
            st.json(data['page_info'])
        with st.expander("Forms/Fields"):
            st.json(data['forms'])
        with st.expander("Extra Buttons (outside forms)"):
            st.json(data['extra_buttons'])
        with st.expander("Interactive Links"):
            st.json(data['interactive_links'])
        with st.expander("Alert/Error Messages"):
            st.json(data['alert_messages'])
    else:
        st.info("Please start by scraping a website.")

# ---------------- TAB 2: USER STORY ----------------------------------------
with tab[1]:
    st.subheader("User Story Creation Agent")
    if 'scraped_data' not in st.session_state:
        st.warning("Please scrape a website first in Tab 1.")
    else:
        if st.button("Generate User Stories (AI Agent)", key="story-gen-btn"):
            with st.spinner("LLM agent writing user stories..."):
                user_story_md = generate_user_stories(st.session_state['scraped_data'])
                st.session_state['user_story_md'] = user_story_md
                story_table = parse_markdown_user_stories(user_story_md)
                st.session_state['user_story_struct'] = story_table
                excel_bytes = style_user_stories_excel(story_table)
                st.session_state['user_story_excel'] = excel_bytes
        if 'user_story_md' in st.session_state:
            st.markdown("#### 📋 User Stories (Guest User):")
            st.markdown(st.session_state['user_story_md'])
            st.markdown("#### Excel Preview:")
            st.dataframe(st.session_state['user_story_struct'])
            st.download_button("Download Excel (User Stories)",
                               data=st.session_state['user_story_excel'],
                               file_name="UserStories.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ---------------- TAB 3: TEST PLAN AGENT -----------------------------------
with tab[2]:
    st.subheader("Test Plan Creation Agent")
    if 'scraped_data' not in st.session_state:
        st.warning("Please scrape a website first in Tab 1.")
    else:
        if st.button("Generate Test Plan (AI Agent)", key="plan-gen-btn"):
            with st.spinner("LLM agent writing test plan..."):
                plan_md = generate_test_plan(st.session_state['scraped_data'])
                st.session_state['test_plan_md'] = plan_md
        if 'test_plan_md' in st.session_state:
            st.markdown("#### 📝 Test Plan (Guest User):")
            st.markdown(st.session_state['test_plan_md'])

# ---------------- TAB 4: TEST CASE + TEST DATA --------------------------
with tab[3]:
    st.subheader("Test Case Development Agent (+ Test Data Generator)")
    if 'scraped_data' not in st.session_state:
        st.warning("Please scrape a website first in Tab 1.")
    else:
        run = st.button("Generate Test Cases/Sheets (AI Agent)", key="tc-gen-btn")
        if run:
            with st.spinner("Generating test cases..."):
                tc_md = generate_test_cases(st.session_state['scraped_data'])
                st.session_state['test_case_md'] = tc_md
                test_cases = parse_markdown_test_cases(tc_md)
                user_story_list = st.session_state.get('user_story_struct', [])
                for i, tc in enumerate(test_cases):
                    if user_story_list:
                        tc['User Story'] = user_story_list[i % len(user_story_list)]['ID']
                    else:
                        tc['User Story'] = 'US-000'
                for tc in test_cases:
                    test_data = generate_test_data_for_case(tc['Test Steps'] + '\n' + tc['Test Scenario'])
                    if re.search(r"\bN/?A\b", test_data, flags=re.I):
                        test_data = "N/A"
                    else:
                        test_data = re.sub(r'^.*response is:?\s*', '', test_data, flags=re.I).strip().strip("'\"")
                        if not test_data or test_data.lower() in ['none', '-', '']: 
                            test_data = "N/A"
                    tc['Test Data'] = test_data
                test_cases = assign_test_case_ids(test_cases)
                excel_bytes = style_test_case_excel(test_cases)
                st.session_state['test_case_excel'] = excel_bytes
                st.session_state['test_case_struct'] = test_cases
        if 'test_case_md' in st.session_state:
            st.markdown("#### 🧾 Test Cases (Guest User):")
            st.markdown(st.session_state['test_case_md'])
        if 'test_case_struct' in st.session_state:
            st.markdown("#### Excel Preview:")
            st.dataframe(st.session_state['test_case_struct'])
            st.download_button("Download Excel (Test Cases + Data)",
                               data=st.session_state['test_case_excel'],
                               file_name="TestCases.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ---------------- TAB 5: BROWSER TEST AGENT --------------------------
with tab[4]:
    st.subheader("Browser Test Agent")
    if 'scraped_data' not in st.session_state:
        st.warning("Please scrape a website in Tab 1 first.")
    else:
        st.info("This agent will use the test cases generated in Tab 4. Make sure to generate test cases first.")
        
        # Option to upload JSON file or use generated test cases
        st.markdown("#### Test Case Source:")
        test_source = st.radio(
            "Choose test case source:",
            ["Use Generated Test Cases (from Tab 4)", "Upload JSON File"],
            key="test_source"
        )
        
        json_test_cases = []
        
        if test_source == "Upload JSON File":
            uploaded_file = st.file_uploader("Upload JSON test cases file", type="json")
            if uploaded_file:
                try:
                    json_test_cases = json.load(uploaded_file)
                    st.success(f"Loaded {len(json_test_cases)} test cases from uploaded file.")
                except Exception as e:
                    st.error(f"Error loading JSON file: {str(e)}")
        else:
            # Use generated test cases from Tab 4
            if 'test_case_struct' in st.session_state:
                json_test_cases = st.session_state['test_case_struct']
                st.success(f"Using {len(json_test_cases)} test cases generated in Tab 4.")
            else:
                st.warning("No test cases found. Please generate test cases in Tab 4 first.")
        
        if st.button("Run Browser Test Agent", key="browser-test-btn"):
            if not json_test_cases:
                st.error("No test cases available. Please generate test cases in Tab 4 or upload a JSON file.")
            else:
                with st.spinner("Running browser test agent step by step..."):
                    url = st.session_state['scraped_data']['url']
                    scraped_data = st.session_state['scraped_data']
                    
                    # Create new event loop for async execution
                    loop = asyncio.new_event_loop()
                    asyncio.set_event_loop(loop)
                    try:
                        test_results, button_report, error_occurred = loop.run_until_complete(
                            run_browser_test_agent(json_test_cases, url, scraped_data)
                        )
                        
                        st.session_state['browser_test_results'] = test_results
                        st.session_state['button_report'] = button_report
                        st.session_state['browser_test_status'] = "Success" if not error_occurred else "Completed with errors"
                        
                        # Generate Excel reports
                        if button_report:
                            excel_bytes = style_button_report_excel(button_report)
                            st.session_state['button_report_excel'] = excel_bytes

                        # Generate test execution report
                        report_data = []
                        for result in test_results:
                            test_id = result.get('test_id', 'UNKNOWN')
                            test_scenario = result.get('test_scenario', '')
                            
                            # Build steps summary
                            steps_summary = []
                            actual_results = []
                            overall_status = result.get('overall_status', 'UNKNOWN')
                            
                            if 'error' in result:
                                actual_results.append(f"Execution error: {result['error']}")
                                overall_status = "ERROR"
                            elif 'steps' in result:
                                for step in result['steps']:
                                    action_desc = step.get('description', step.get('action', {}).get('action', 'Unknown action'))
                                    steps_summary.append(action_desc)
                                    actual_results.append(f"{action_desc}: {'PASS' if step.get('result') else 'FAIL'}")
                            
                            # Find corresponding test case for expected result
                            expected_result = "N/A"
                            for tc in json_test_cases:
                                if tc.get('Test Case ID') == test_id:
                                    expected_result = tc.get('Expected Result', 'N/A')
                                    break
                            
                            report_data.append({
                                "Test Case ID": test_id,
                                "Test Scenario": test_scenario,
                                "Test Steps": "\n".join(steps_summary),
                                "Expected Result": expected_result,
                                "Actual Result": "\n".join(actual_results),
                                "Status": overall_status,
                                "Execution Timestamp": datetime.now().isoformat()
                            })
                        
                        st.session_state['test_execution_report'] = report_data
                        excel_bytes = style_test_execution_report_excel(report_data)
                        st.session_state['test_execution_excel'] = excel_bytes
                        
                    except Exception as e:
                        st.error(f"Browser test execution failed: {str(e)}")
                    finally:
                        loop.close()

        # Display results if available
        if 'browser_test_results' in st.session_state:
            st.markdown("#### 🖥️ Browser Test Results:")
            
            # Summary
            total_tests = len(st.session_state['browser_test_results'])
            passed_tests = sum(1 for result in st.session_state['browser_test_results'] 
                             if result.get('overall_status') == 'PASS')
            failed_tests = total_tests - passed_tests
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total Tests", total_tests)
            with col2:
                st.metric("Passed", passed_tests)
            with col3:
                st.metric("Failed", failed_tests)
            
            # Detailed results
            for result in st.session_state['browser_test_results']:
                with st.expander(f"Test ID: {result.get('test_id', 'UNKNOWN')} - {result.get('overall_status', 'UNKNOWN')}"):
                    st.write(f"**Test Scenario:** {result.get('test_scenario', 'N/A')}")
                    
                    if 'error' in result:
                        st.error(f"Execution Error: {result['error']}")
                    else:
                        st.write("**Step Results:**")
                        for i, step in enumerate(result.get('steps', []), 1):
                            action_info = step.get('action', {})
                            result_status = "✅ PASS" if step.get('result') else "❌ FAIL"
                            description = step.get('description', action_info.get('action', 'Unknown'))
                            st.write(f"{i}. {description} - {result_status}")
                            
                            if 'error' in step:
                                st.error(f"   Error: {step['error']}")
            
            # Download buttons for reports
            if 'test_execution_report' in st.session_state:
                st.markdown("#### 📊 Test Execution Report:")
                st.dataframe(st.session_state['test_execution_report'])
                st.download_button(
                    "Download Excel (Test Execution Results)",
                    data=st.session_state['test_execution_excel'],
                    file_name=f"TestExecutionResults_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
            if 'button_report' in st.session_state:
                st.markdown("#### 🔘 Button Test Report:")
                st.dataframe(st.session_state['button_report'])
                st.download_button(
                    "Download Excel (Button Test Report)",
                    data=st.session_state['button_report_excel'],
                    file_name=f"ButtonTestReport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

st.caption("Uses Playwright, BeautifulSoup, LangChain, Llama-3 (Groq), and openpyxl. Integrated test case generation and browser automation.")
